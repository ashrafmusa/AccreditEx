#!/usr/bin/env python3
"""
Extract Oman Health Accreditation Standards (SMCS) from Excel files.

Reads all SMCS*.xlsx files from the docs folder and produces:
  1. standards_smcs_import.json  — ready to upload to Firestore 'standards' collection
  2. program_update.json         — updated program metadata

Output matches the AccreditEx system schema:
  Standard { id, standardId, programId, chapter, section, description, criticality, totalMeasures, subStandards[] }
"""

import json
import os
import re
from datetime import datetime, timezone

import openpyxl

DOCS_DIR = r"D:\_Projects\accreditex\docs\Oman Health Accreditation Standerds"
OUTPUT_DIR = r"D:\_Projects\accreditex\data\sample-data"
PROGRAM_ID = "prog-ohap"
CHAPTER = "Chapter 4"
SECTION = "Specialized Medical Care Services (SMCS)"


def extract_standards_from_workbook(filepath: str) -> list[dict]:
    """Parse one SMCS Excel workbook and return a list of Standard objects."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    standards: list[dict] = []
    current_standard: dict | None = None
    current_subsection: str | None = None  # e.g. "Provision of Care", "Burn Care Unit"

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cell_a = str(row[0].value or "").strip()
        cell_b = str(row[1].value or "").strip() if len(row) > 1 else ""
        cell_c = str(row[2].value or "").strip() if len(row) > 2 else ""

        # Skip header rows
        if not cell_a or cell_a in ("Quality Assurance Center", "Oman Healthcare Accrediation System",
                                     "Self Assessment Tool (SAT)", "Chapter 4"):
            continue

        # Detect sub-section headings (no dot, not a standard ID, cell B is empty or same as A)
        # These are category labels like "Provision of Care", "Burn Care Unit", etc.
        smcs_match = re.match(r'^(?:SMCS|MSCS)\.?(\d+)\s*$', cell_a)
        sub_match = re.match(r'^(?:SMCS|MSCS)\.?(\d+)\.([a-z])\s*$', cell_a, re.IGNORECASE)

        if smcs_match and cell_b:
            # This is a main standard row: SMCS.1, SMCS.98, etc.
            # Save previous standard
            if current_standard:
                standards.append(current_standard)

            std_num = smcs_match.group(1)
            standard_id = f"OHAS.SMCS.{std_num}"

            # Parse total measures from cell C  e.g. "Total Measures: 3"
            measures_match = re.search(r'Total Measures:\s*(\d+)', cell_c)
            total_measures = int(measures_match.group(1)) if measures_match else 0

            # Determine criticality based on context
            criticality = "High" if total_measures >= 6 else "Medium"

            current_standard = {
                "id": standard_id,
                "standardId": standard_id,
                "programId": PROGRAM_ID,
                "chapter": CHAPTER,
                "section": SECTION,
                "description": cell_b,
                "criticality": criticality,
                "totalMeasures": total_measures,
                "subStandards": [],
            }
            if current_subsection:
                current_standard["category"] = current_subsection

        elif sub_match and cell_b:
            # This is a sub-standard row: SMCS.1.a, SMCS.98.b, etc.
            std_num = sub_match.group(1)
            sub_letter = sub_match.group(2).upper()
            sub_id = f"OHAS.SMCS.{std_num}.{sub_letter}"

            if current_standard:
                current_standard["subStandards"].append({
                    "id": sub_id,
                    "description": cell_b,
                })

        elif not smcs_match and not sub_match and cell_a and not cell_b:
            # Likely a sub-section heading (e.g. "Provision of Care", "Burn Care Unit")
            # Skip rows that look like scoring headers
            if "Conformity" not in cell_a and "Total Score" not in cell_a and "Quotation" not in cell_a:
                current_subsection = cell_a

    # Don't forget the last standard
    if current_standard:
        standards.append(current_standard)

    wb.close()
    return standards


def main():
    all_standards: list[dict] = []

    # Get all SMCS xlsx files sorted by number
    xlsx_files = sorted(
        [f for f in os.listdir(DOCS_DIR) if f.lower().endswith(".xlsx") and f.upper().startswith("SMCS")],
        key=lambda f: int(re.search(r'(\d+)', f).group(1)) if re.search(r'(\d+)', f) else 0,
    )

    print(f"Found {len(xlsx_files)} SMCS Excel files")
    print("=" * 60)

    for fname in xlsx_files:
        filepath = os.path.join(DOCS_DIR, fname)
        standards = extract_standards_from_workbook(filepath)
        print(f"  {fname}: {len(standards)} standards extracted")
        for s in standards:
            sub_count = len(s["subStandards"])
            print(f"    {s['standardId']}: {s['description'][:70]}... ({sub_count} sub-standards)")
        all_standards.extend(standards)

    print("=" * 60)
    print(f"TOTAL: {len(all_standards)} standards with "
          f"{sum(len(s['subStandards']) for s in all_standards)} sub-standards")

    # Save as plain array (compatible with Standards Import modal)
    out_path = os.path.join(OUTPUT_DIR, "standards_smcs_import.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_standards, f, indent=2, ensure_ascii=False)

    print(f"\n✅ Saved to: {out_path}")


if __name__ == "__main__":
    main()
